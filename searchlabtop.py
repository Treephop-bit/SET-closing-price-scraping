from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
import time

from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment

#กำหนด path ของ chrome driver exe
path = r'C:\Users\ken25\OneDrive\Desktop\01_Python\web_scrabling\chromedriver.exe'
service = Service(path)

driver = webdriver.Chrome(service=service)

#เปิดเว็บไซต์
web_url = 'https://www.central.co.th/th/electronic-gadgets/computers/laptops'
driver.get(web_url)

#หา elements
products = driver.find_elements(By.CLASS_NAME, 'flip-card-front')

brandlist = []
modellist = []
urllist = []

for p in products:
    listcheck = p.text.split('\n')
    #print(listcheck)
    brand = listcheck[-3]
    model = listcheck[-2]
    #print(brand)
    #print(model)
    
    links = p.find_element(By.CLASS_NAME, 'sliderText')
    alltag = links.find_elements(By.TAG_NAME, 'a')
    url = alltag[-1].get_attribute('href')
    #print(url)
    #print('-----------------------------------------')
    brandlist.append(brand)
    modellist.append(model)
    urllist.append(url)
    
print(brandlist)
print(modellist)
print(urllist)

# write and save files
wb = Workbook()
ws = wb.active #บอกให้รู้ว่า worksheet ปัจจุบันคือตัวไหน

ws['A1'] = 'Brand'
ws['A1'].font = Font(bold=True)
ws['A1'].alignment = Alignment(horizontal='center')
ws['B1'] = 'Model'
ws['B1'].font = Font(bold=True)
ws['B1'].alignment = Alignment(horizontal='center')
ws['C1'] = 'URL'
ws['C1'].font = Font(bold=True)
ws['C1'].alignment = Alignment(horizontal='center')



ws.column_dimensions['B'].width = 70
ws.column_dimensions['C'].width = 70

for v in zip(brandlist, modellist, urllist):
    ws.append(v)


wb.save('searchlabtop.xlsx')



time.sleep(5)
driver.quit() #.close จะปิด web browser แต่ driver ยังทำงานอยู่ แต่ quit จะปิด driver ด้วย ควรใช้ .quit